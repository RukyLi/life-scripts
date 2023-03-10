import os
import time
from datetime import datetime
import openpyxl

from bs4 import BeautifulSoup
import requests

headers = {
    'Accept': 'application/json, text/plain, */*',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36',
    'Proxy-Connection': 'keep-alive'
}

code_and_org_id_url = "http://www.cninfo.com.cn/new/information/topSearch/query?keyWord="
hisAnnouncement_url = "http://www.cninfo.com.cn/new/hisAnnouncement/query"

find_key = ["2018年年度报告", "2019年年度报告", "2020年年度报告", "2021年年度报告", "2022年年度报告"]
download_ref_url = "http://www.cninfo.com.cn/new/announcement/bulletin_detail?announceId={}&flag=true&announceTime={}"
data_ref_url = "http://www.cninfo.com.cn/data20/financialData/getMainIndicators?scode={}&sign=1"
data_ref_balance_url = "http://www.cninfo.com.cn/data20/financialData/getBalanceSheets?scode={}&sign=1"

base_download_url = "D:\data\company\\"

choice_num = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']


def get_org_id(num):
    response_content = requests.post(code_and_org_id_url + num, headers)
    content_json = response_content.json()
    if len(content_json) == 0:
        return None;
    else:
        return content_json[0].get("orgId")


def get_hisAnnouncement(num, code):
    data_query = str(num) + "," + str(code)
    query = {
        "stock": data_query,
        "tabName": "fulltext",
        "pageSize": 30,
        "pageNum": 1,
        "column": "szse",
        "category": "category_ndbg_szsh;",
        "plate": "sz",
        "isHLtitle": True
    }
    response_content = requests.post(url=hisAnnouncement_url, data=query, headers=headers)
    data = response_content.json().get("announcements")
    if data is None:
        query.update({"column": "sse"})
        query.update({"plate": "sh"})
        response_content = requests.post(url=hisAnnouncement_url, data=query, headers=headers)
    data = response_content.json().get("announcements")
    if data is None:
        return None
    else:
        return data


def filter_with_time(hisAnnouncement):
    fit_data = []
    # print(hisAnnouncement)
    for history in hisAnnouncement:
        title = history.get("announcementTitle")
        for key in find_key:
            if key in title:
                fit_data.append(history)
    return fit_data


def analysis_history(fit_data):
    download_url = {}
    for data in fit_data:
        id = data.get("announcementId")
        time = datetime.fromtimestamp(data.get("announcementTime") / 1000).strftime('%Y-%m-%d')
        url = download_ref_url.format(str(id), time)
        url_content = requests.post(url=url, headers=headers)
        download_url[data.get("announcementTitle")] = url_content.json().get("fileUrl")
    return download_url


def start_download(url_list, dir_name):
    dir = base_download_url + dir_name
    if not os.path.exists(dir):
        os.mkdir(dir)
    else:
        print("文件夹以存在，跳过下载")
        return
    for key, value in url_list.items():
        response = requests.get(value)
        print("下载文件：" + dir + "\\" + key.strip() + ".PDF")
        with open(dir + "\\" + key.strip() + ".PDF", 'wb') as f:
            f.write(response.content)


def find_data(num, name):
    hisAnnouncement = get_hisAnnouncement(str(num), get_org_id(num))
    if hisAnnouncement is None:
        print("未找到任何符合条件的pdf文件")
        return
    fit_data = filter_with_time(hisAnnouncement)
    download_url_list = analysis_history(fit_data)
    start_download(download_url_list, name)


def create_excel_data(code, dir_name):
    url = data_ref_url.format(str(code))
    response = requests.get(url, headers=headers).json()
    rate_of_return_response = response.get("data").get("records")[0].get("year")

    balance_url = data_ref_balance_url.format(str(code))
    balance_response = requests.get(balance_url, headers=headers).json()
    balance_data = balance_response.get("data").get("records")[0].get("year")

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = "总资产报酬率"

    index = 0
    for year in rate_of_return_response:
        worksheet[choice_num[index] + "1"] = year.get("ENDDATE")
        worksheet[choice_num[index] + "2"] = str(year.get("F016N"))
        index = index + 1

    title1 = balance_data[2].pop('index')
    # print(title1)
    worksheet['A4'] = title1
    sorted_1 = dict(sorted(balance_data[2].items(), key=lambda x: x[0]))
    index = 0
    for key, value in sorted_1.items():
        worksheet[choice_num[index] + "4"] = str(key)
        worksheet[choice_num[index] + "5"] = str(value)
        index = index + 1

    title2 = balance_data[3].pop('index')
    # print(title2)
    worksheet['A7'] = title2
    sorted_2 = dict(sorted(balance_data[3].items(), key=lambda x: x[0]))
    index = 0
    for key, value in sorted_2.items():
        worksheet[choice_num[index] + "7"] = str(key)
        worksheet[choice_num[index] + "8"] = str(value)
        index = index + 1

    title3 = balance_data[4].pop('index')
    # print(title3)
    worksheet['A10'] = title3
    sorted_3 = dict(sorted(balance_data[4].items(), key=lambda x: x[0]))
    index = 0
    for key, value in sorted_3.items():
        worksheet[choice_num[index] + "10"] = str(key)
        worksheet[choice_num[index] + "11"] = str(value)
        index = index + 1

    dir = base_download_url + dir_name
    if not os.path.exists(dir):
        os.mkdir(dir)
    dir_excel = dir + "\\" + "数据表.xlsx"
    if os.path.exists(dir_excel):
        os.remove(dir_excel)
    workbook.save(dir_excel)


def read_company(dir):
    workbook = openpyxl.load_workbook(dir)
    worksheet = workbook.active

    max_row = worksheet.max_row
    for row in range(max_row):
        code = worksheet.cell(row=row + 1, column=1).value.strip()
        company = worksheet.cell(row=row + 1, column=2).value.strip()
        print("开始爬取：" + company)
        find_data(code, company)
        create_excel_data(code, company)
        print("爬取结束")
        time.sleep(5)


def only(code, name):
    print("开始爬取：" + name)
    find_data(code, name)
    create_excel_data(code, name)
    print("爬取结束")


if __name__ == '__main__':
    # read_company("D:\data\company\\kk.xlsx")
    only("601877", "正泰电器")
